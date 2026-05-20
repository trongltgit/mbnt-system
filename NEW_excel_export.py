"""
📊 Module xuất báo cáo
- Export Excel (.xlsx)
- Export CSV (.csv)  
- Export PDF (.pdf)
- Phân quyền: PNV thấy phòng mình, PQL thấy tất cả
- PNV không thấy tỷ giá hệ thống
"""

import io
import csv
from datetime import datetime
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from models import Transaction, Message, TransactionEdit, User
from permissions import PermissionManager

class ExcelExporter:
    """Xuất dữ liệu sang Excel"""
    
    @staticmethod
    def to_xlsx(transactions: List[Transaction], 
                messages: List[Message],
                edits: List[TransactionEdit],
                user_role: str,
                current_user_id: int = None,
                current_user_dept: str = None) -> bytes:
        """
        Xuất Excel với multiple sheets
        """
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Xóa sheet mặc định
        
        # 1. Sheet TRANSACTIONS
        ExcelExporter._create_transaction_sheet(
            wb, transactions, user_role, current_user_id, current_user_dept, "TRANSACTIONS"
        )
        
        # 2. Sheet MESSAGES
        ExcelExporter._create_message_sheet(wb, messages, "MESSAGES")
        
        # 3. Sheet EDITS (chỉ PQL xem)
        if user_role == "pql":
            ExcelExporter._create_edit_sheet(wb, edits, "EDIT_LOG")
        
        # 4. Sheet SUMMARY (Tổng hợp)
        ExcelExporter._create_summary_sheet(wb, transactions, "SUMMARY")
        
        # Lưu vào memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def _create_transaction_sheet(wb, transactions, user_role, current_user_id, current_user_dept, sheet_name):
        """Tạo sheet giao dịch"""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = [
            "Transaction No", "Date", "Customer", "Direction", "Currency",
            "Amount", "Buy Rate", "Sell Rate", "Profit", "Status", "PNV"
        ]
        
        # Thêm system_rate columns nếu PQL/BGD/PGD xem
        if PermissionManager.can_see_system_rate(user_role):
            headers.insert(6, "System Buy Rate")
            headers.insert(7, "System Sell Rate")
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for txn in transactions:
            # Kiểm tra quyền xem
            if user_role == "pnv" and current_user_id != txn.pnv_id:
                continue
            if user_role == "pgd" and current_user_dept != txn.pnv.department:
                continue
            
            row = [
                txn.transaction_no,
                txn.transaction_date.strftime("%Y-%m-%d %H:%M") if txn.transaction_date else "",
                txn.customer.customer_name if txn.customer else "",
                txn.direction.upper(),
                txn.currency_code,
                txn.amount,
                txn.buy_rate or "",
                txn.sell_rate or "",
                txn.profit or "",
                txn.status,
                txn.pnv.full_name if txn.pnv else ""
            ]
            
            # Insert system rates nếu có quyền
            if PermissionManager.can_see_system_rate(user_role):
                row.insert(6, txn.system_buy_rate or "")
                row.insert(7, txn.system_sell_rate or "")
            
            ws.append(row)
        
        # Auto width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_message_sheet(wb, messages, sheet_name):
        """Tạo sheet tin nhắn"""
        ws = wb.create_sheet(sheet_name)
        
        headers = ["Date", "From", "Department", "To", "Message Type", "Content", "Read", "Answered"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for msg in messages:
            ws.append([
                msg.created_at.strftime("%Y-%m-%d %H:%M") if msg.created_at else "",
                msg.sender.full_name if msg.sender else "",
                msg.sender.department if msg.sender else "",
                msg.recipient_name or msg.recipient_department or "",
                msg.message_type.upper(),
                msg.content[:100] + "..." if len(msg.content) > 100 else msg.content,
                "Yes" if msg.is_read else "No",
                "Yes" if msg.is_answered else "No"
            ])
        
        # Auto width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @staticmethod
    def _create_edit_sheet(wb, edits, sheet_name):
        """Tạo sheet lịch sử chỉnh sửa (chỉ PQL xem)"""
        ws = wb.create_sheet(sheet_name)
        
        headers = ["Date", "Transaction", "Edited By", "Field", "Old Value", "New Value", "Reason", "Type"]
        ws.append(headers)
        
        # Style headers  
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for edit in edits:
            ws.append([
                edit.edited_at.strftime("%Y-%m-%d %H:%M") if edit.edited_at else "",
                edit.transaction.transaction_no if edit.transaction else "",
                edit.edited_user.full_name if edit.edited_user else "",
                edit.field_name,
                edit.old_value or "",
                edit.new_value or "",
                edit.reason or "",
                edit.change_type or ""
            ])
        
        # Auto width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @staticmethod
    def _create_summary_sheet(wb, transactions, sheet_name):
        """Tạo sheet tổng hợp"""
        ws = wb.create_sheet(sheet_name)
        
        # Statistics
        total_txn = len(transactions)
        buy_txn = len([t for t in transactions if t.direction == "buy"])
        sell_txn = len([t for t in transactions if t.direction == "sell"])
        
        total_profit = sum([t.profit or 0 for t in transactions])
        
        ws.append(["SUMMARY REPORT"])
        ws.append([""])
        ws.append(["Metric", "Value"])
        ws.append(["Total Transactions", total_txn])
        ws.append(["Buy Transactions", buy_txn])
        ws.append(["Sell Transactions", sell_txn])
        ws.append(["Total Profit", total_profit])
        ws.append(["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
        
        # Style title
        ws['A1'].font = Font(bold=True, size=14)
        
        # Style header
        for cell in ws[3]:
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.font = Font(bold=True)


class CSVExporter:
    """Xuất dữ liệu sang CSV"""
    
    @staticmethod
    def to_csv(transactions: List[Transaction], 
               user_role: str,
               current_user_id: int = None,
               current_user_dept: str = None) -> bytes:
        """Xuất CSV"""
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "Transaction_No", "Date", "Customer", "Direction", "Currency",
            "Amount", "Buy_Rate", "Sell_Rate", "Profit", "Status", "PNV"
        ]
        
        if PermissionManager.can_see_system_rate(user_role):
            headers.insert(6, "System_Buy_Rate")
            headers.insert(7, "System_Sell_Rate")
        
        writer.writerow(headers)
        
        # Data rows
        for txn in transactions:
            # Check permissions
            if user_role == "pnv" and current_user_id != txn.pnv_id:
                continue
            if user_role == "pgd" and current_user_dept != txn.pnv.department:
                continue
            
            row = [
                txn.transaction_no,
                txn.transaction_date.strftime("%Y-%m-%d %H:%M") if txn.transaction_date else "",
                txn.customer.customer_name if txn.customer else "",
                txn.direction.upper(),
                txn.currency_code,
                txn.amount,
                txn.buy_rate or "",
                txn.sell_rate or "",
                txn.profit or "",
                txn.status,
                txn.pnv.full_name if txn.pnv else ""
            ]
            
            if PermissionManager.can_see_system_rate(user_role):
                row.insert(6, txn.system_buy_rate or "")
                row.insert(7, txn.system_sell_rate or "")
            
            writer.writerow(row)
        
        return output.getvalue().encode('utf-8')


class PDFExporter:
    """Xuất dữ liệu sang PDF"""
    
    @staticmethod
    def to_pdf(transactions: List[Transaction],
               user_role: str,
               current_user_id: int = None,
               current_user_dept: str = None) -> bytes:
        """Xuất PDF"""
        
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not installed. Install with: pip install reportlab")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=10*mm)
        
        # Build table data
        data = [[
            "Transaction No", "Date", "Customer", "Direction", "Currency",
            "Amount", "Buy Rate", "Sell Rate", "Profit", "Status"
        ]]
        
        for txn in transactions:
            if user_role == "pnv" and current_user_id != txn.pnv_id:
                continue
            if user_role == "pgd" and current_user_dept != txn.pnv.department:
                continue
            
            data.append([
                txn.transaction_no,
                txn.transaction_date.strftime("%Y-%m-%d") if txn.transaction_date else "",
                txn.customer.customer_name if txn.customer else "",
                txn.direction.upper(),
                txn.currency_code,
                f"{txn.amount:,.0f}" if txn.amount else "",
                f"{txn.buy_rate:,.4f}" if txn.buy_rate else "",
                f"{txn.sell_rate:,.4f}" if txn.sell_rate else "",
                f"{txn.profit:,.0f}" if txn.profit else "",
                txn.status
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        # Build PDF
        elements = [table]
        doc.build(elements)
        
        output.seek(0)
        return output.getvalue()
